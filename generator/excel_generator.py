"""
Excel-Dokument-Generator
Ungeverändert aus WBI-Docu-Assist übernommen.
"""

import io
import re
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _safe_filename(title: str) -> str:
    return re.sub(r'[^\w\s\-äöüÄÖÜß]', '_', title).strip() + ".xlsx"


def _apply_header(cell):
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F3864')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def generate_excel(data: dict, template_path: str):
    title = data.get('titleSubject', '') + ' - ' + data.get('titleTopic', '')
    title = title.strip(' -')
    template_id = data.get('template', 'netzwerk')
    chapters = data.get('chapters', [])
    refs = [r for r in data.get('refs', []) if r.get('num') or r.get('name')]
    markdown_content = data.get('markdownContent', '')

    wb = load_workbook(template_path)

    if template_id == 'netzwerk':
        _fill_netzwerk(wb, data, title)
    elif template_id == 'brief':
        _fill_brief(wb, data, title)
    else:
        _fill_generic(wb, data, title, chapters, refs, markdown_content)

    buf = io.BytesIO()
    wb.save(buf)
    return buf, _safe_filename(title)


def _fill_netzwerk(wb, data, title):
    today = datetime.now().strftime('%d.%m.%Y')
    subject = data.get('titleSubject', '')
    if 'Info' in wb.sheetnames:
        ws = wb['Info']
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Kundenname':
                    ws.cell(row=cell.row, column=cell.column + 1, value=subject)
                elif cell.value == 'Netzwerkdoku':
                    ws.cell(row=cell.row, column=cell.column + 1, value=title)
                elif str(cell.value or '').startswith('Erstellt:'):
                    ws.cell(row=cell.row, column=cell.column + 1, value=today)


def _fill_brief(wb, data, title):
    ws = wb.active
    today = datetime.now().strftime('%d.%m.%Y')
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').lower() in ('datum', 'date'):
                ws.cell(row=cell.row, column=cell.column + 1, value=today)
            if str(cell.value or '').lower() in ('betreff', 'subject'):
                ws.cell(row=cell.row, column=cell.column + 1, value=title)


def _fill_generic(wb, data, title, chapters, refs, markdown):
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    row_num = 1
    ws.cell(row=row_num, column=1, value=title)
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=16, color='1F3864')
    ws.row_dimensions[row_num].height = 30
    row_num += 2
    ws.cell(row=row_num, column=1, value='Erstellt:')
    ws.cell(row=row_num, column=2, value=datetime.now().strftime('%d.%m.%Y'))
    row_num += 2
    if refs:
        ws.cell(row=row_num, column=1, value='DMS-Link / WBI-Nummer')
        ws.cell(row=row_num, column=2, value='Bezeichnung')
        _apply_header(ws.cell(row=row_num, column=1))
        _apply_header(ws.cell(row=row_num, column=2))
        row_num += 1
        for ref in refs:
            ws.cell(row=row_num, column=1, value=ref.get('num', ''))
            ws.cell(row=row_num, column=2, value=ref.get('name', ''))
            row_num += 1
        row_num += 1
    if markdown.strip():
        for line in markdown.split('\n'):
            stripped = line.strip()
            if not stripped or stripped == '---':
                row_num += 1; continue
            if stripped.startswith('# '):
                ws.cell(row=row_num, column=1, value=stripped[2:]).font = Font(bold=True, size=14, color='1F3864')
            elif stripped.startswith('## '):
                ws.cell(row=row_num, column=1, value=re.sub(r'^\d+\.\s*', '', stripped[3:])).font = Font(bold=True, size=12, color='2E75B6')
            elif stripped.startswith('### '):
                ws.cell(row=row_num, column=1, value='    ' + re.sub(r'^[\d.]+\s*', '', stripped[4:])).font = Font(bold=True, size=11)
            elif stripped.startswith('|') and not re.match(r'^\|[-| :]+\|$', stripped):
                cells = [c.strip() for c in stripped.strip('|').split('|')]
                for ci, val in enumerate(cells, 1):
                    ws.cell(row=row_num, column=ci, value=val)
            elif not stripped.startswith(('>', '-', '*', '`', '#', '|', '!')):
                ws.cell(row=row_num, column=1, value=stripped)
            row_num += 1
    else:
        valid_chapters = [c for c in chapters if c.get('name', '').strip()]
        for idx, ch in enumerate(valid_chapters, 1):
            ws.cell(row=row_num, column=1, value=f'{idx}. {ch["name"]}').font = Font(bold=True, size=12, color='1F3864')
            row_num += 1
            for si, sub in enumerate([s for s in ch.get('subs', []) if s.strip()], 1):
                ws.cell(row=row_num, column=1, value=f'    {idx}.{si} {sub}')
                row_num += 1
            row_num += 1
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
